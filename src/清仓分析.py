# -*- coding: utf-8 -*-
"""
清仓分析脚本 - 新浪行情版

功能：
1. 读取统一配置中的清仓分析表
2. 使用新浪实时行情接口获取当前价格
3. 使用腾讯日K接口获取近200日最高价
4. 对比清仓价格，计算清仓后涨幅
5. 计算现价相比近200日最高价跌幅
6. 输出新的 Excel 文件
7. 不改变原文件顺序
8. 输出文件名带日期 + 时间，避免 PermissionError

依赖：
pip install pandas openpyxl requests
"""

import argparse
import os
import time
import json
import logging
from datetime import datetime

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from logging_utils import configure_root_rotating_logging, default_log_dir
from pipeline_config import config_value, project_path, resolve_input
from stock_utils import market_suffix, normalize_code_digits as normalize_code, timestamped_output_path


# =========================
# 配置区
# =========================

SHEET_NAME = 0

CODE_COL = "代码"
NAME_COL = "名称"
CLEAR_PRICE_COL = "清仓价格"
CURRENT_PRICE_COL = "现价"
AFTER_CLEAR_GAIN_COL = "清仓后涨幅"

HIGH_200_COL = "近200日最高价"
HIGH_200_DRAWDOWN_COL = "现价相比近200日最高价跌幅"

OUTPUT_FILE = timestamped_output_path(
    project_path(config_value("files", "output_dir", "data/output")),
    "清仓分析_清仓后涨幅",
)

REQUEST_TIMEOUT = 10
RETRY_TIMES = 3
SLEEP_SECONDS = 1


def sina_symbol(code):
    """
    新浪行情代码格式：
    深市：sz002536
    沪市：sh601133
    北交所：bjxxxxxx，新浪不一定稳定支持
    """
    code = normalize_code(code)

    if not code:
        return ""

    return market_suffix(code).lower() + code


def create_session(use_system_proxy=True):
    """
    创建 requests session。

    use_system_proxy=True:
        使用你当前系统代理。

    use_system_proxy=False:
        不读取系统代理。

    这版会先用系统代理请求，如果失败，再自动改成不走系统代理。
    """
    session = requests.Session()

    if use_system_proxy:
        session.trust_env = True
    else:
        session.trust_env = False

    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": "https://finance.sina.com.cn/",
        "Accept": "*/*",
        "Connection": "close",
    })

    return session


def parse_sina_response(text):
    """
    解析新浪返回结果。

    新浪返回类似：
    var hq_str_sz002536="飞龙股份,12.34,12.10,12.50,...";

    字段含义：
    0 名称
    1 今日开盘价
    2 昨日收盘价
    3 当前价格
    """
    result = {}

    if not text:
        return result

    lines = text.strip().split(";")

    for line in lines:
        line = line.strip()

        if not line:
            continue

        if "hq_str_" not in line:
            continue

        try:
            left, right = line.split("=", 1)

            symbol = left.split("hq_str_")[-1].strip()

            content = right.strip().strip('"')

            if not content:
                continue

            parts = content.split(",")

            if len(parts) < 4:
                continue

            name = parts[0].strip()
            current_price = parts[3].strip()

            if current_price in ["", "0", "0.00", "-"]:
                continue

            price = float(current_price)

            code = symbol[-6:]

            result[code] = {
                "name": name,
                "price": price,
                "symbol": symbol,
            }

        except Exception as e:
            logging.warning(f"解析新浪行情失败，原始行：{line}，错误：{repr(e)}")

    return result


def fetch_price_from_sina(codes, use_system_proxy=True):
    """
    从新浪批量获取股票现价。
    """
    clean_codes = []

    for code in codes:
        code = normalize_code(code)
        if code:
            clean_codes.append(code)

    # 去重，仅用于减少请求次数；不影响最终输出顺序
    clean_codes = list(dict.fromkeys(clean_codes))

    if not clean_codes:
        return {}

    session = create_session(use_system_proxy=use_system_proxy)

    result = {}

    # 新浪接口一批不要太大，稳一点
    batch_size = 50

    for start in range(0, len(clean_codes), batch_size):
        batch_codes = clean_codes[start:start + batch_size]

        symbols = [sina_symbol(code) for code in batch_codes]
        symbols = [x for x in symbols if x]

        url = "http://hq.sinajs.cn/list=" + ",".join(symbols)

        last_error = None

        for attempt in range(1, RETRY_TIMES + 1):
            try:
                logging.info(
                    f"新浪行情：正在获取第 {start + 1} - {start + len(batch_codes)} 只股票，"
                    f"第 {attempt} 次尝试，"
                    f"use_system_proxy={use_system_proxy}"
                )

                response = session.get(
                    url,
                    timeout=REQUEST_TIMEOUT
                )

                response.raise_for_status()

                # 新浪常见编码是 gbk
                response.encoding = "gbk"

                batch_result = parse_sina_response(response.text)

                if not batch_result:
                    raise RuntimeError(f"新浪返回为空或解析为空，返回内容前200字符：{response.text[:200]}")

                result.update(batch_result)

                logging.info(
                    f"新浪行情：本批成功 {len(batch_result)} 条，累计成功 {len(result)} 条"
                )

                break

            except Exception as e:
                last_error = e
                logging.warning(
                    f"新浪行情：本批第 {attempt} 次失败：{repr(e)}"
                )
                time.sleep(SLEEP_SECONDS)

        else:
            logging.error(
                f"新浪行情：本批获取失败，股票代码：{batch_codes}，错误：{repr(last_error)}"
            )

        completed = min(start + len(batch_codes), len(clean_codes))
        progress = 5 + round(25 * completed / len(clean_codes))
        logging.info(
            f"进度：{progress}%（现价已处理 {completed} 只，共 {len(clean_codes)} 只）"
        )

    return result


def fetch_price_map(codes):
    """
    获取价格总入口。

    逻辑：
    1. 先使用系统代理请求新浪
    2. 如果失败，再不走系统代理请求新浪
    """
    logging.info("开始获取股票现价，优先使用新浪行情接口")

    price_map = fetch_price_from_sina(codes, use_system_proxy=True)

    if price_map:
        logging.info(f"使用系统代理获取成功，共 {len(price_map)} 条")
        return price_map

    logging.warning("使用系统代理获取失败，改为不读取系统代理再试一次")

    price_map = fetch_price_from_sina(codes, use_system_proxy=False)

    if price_map:
        logging.info(f"不读取系统代理获取成功，共 {len(price_map)} 条")
        return price_map

    logging.error("两种方式均未获取到行情")
    return {}


def parse_tencent_kline_json(text):
    """
    解析腾讯 K 线接口返回。
    """
    if not text:
        return None

    text = text.strip()

    # 兼容 callback(...)
    if text.startswith("jQuery") or text.startswith("callback"):
        start = text.find("(")
        end = text.rfind(")")
        if start >= 0 and end > start:
            text = text[start + 1:end]

    return json.loads(text)


def fetch_200_day_high_one(code, use_system_proxy=True):
    """
    获取单只股票近200日最高价。

    使用腾讯日K接口：
    https://web.ifzq.gtimg.cn/appstock/app/fqkline/get

    返回：
    float 或 None
    """
    code = normalize_code(code)

    if not code:
        return None

    symbol = sina_symbol(code)

    if not symbol:
        return None

    session = create_session(use_system_proxy=use_system_proxy)

    url = "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"

    params = {
        "param": f"{symbol},day,,,200,"
    }

    last_error = None

    for attempt in range(1, RETRY_TIMES + 1):
        try:
            logging.info(
                f"腾讯200日高点：{code} 第 {attempt} 次尝试，use_system_proxy={use_system_proxy}"
            )

            response = session.get(
                url,
                params=params,
                timeout=REQUEST_TIMEOUT
            )

            response.raise_for_status()

            data = parse_tencent_kline_json(response.text)

            stock_data = data.get("data", {}).get(symbol, {})

            kline_list = None

            if "day" in stock_data:
                kline_list = stock_data.get("day")
            elif "qfqday" in stock_data:
                kline_list = stock_data.get("qfqday")
            elif "hfqday" in stock_data:
                kline_list = stock_data.get("hfqday")

            if not kline_list:
                raise RuntimeError(f"腾讯K线为空：{data}")

            highs = []

            for item in kline_list:
                # 常见格式：
                # [日期, 开盘, 收盘, 最高, 最低, 成交量]
                if len(item) >= 4:
                    high_text = item[3]
                    try:
                        high = float(high_text)
                        if high > 0:
                            highs.append(high)
                    except Exception:
                        pass

            if not highs:
                raise RuntimeError(f"未解析到最高价：{kline_list[:3]}")

            return max(highs)

        except Exception as e:
            last_error = e
            logging.warning(
                f"腾讯200日高点：{code} 第 {attempt} 次失败：{repr(e)}"
            )
            time.sleep(SLEEP_SECONDS)

    logging.error(f"腾讯200日高点：{code} 获取失败，错误：{repr(last_error)}")
    return None


def fetch_200_day_high_map(codes):
    """
    批量获取近200日最高价。

    逻辑：
    1. 单只股票先使用系统代理
    2. 如果失败，再不读取系统代理尝试
    """
    result = {}

    clean_codes = []

    for code in codes:
        code = normalize_code(code)
        if code:
            clean_codes.append(code)

    # 去重，仅用于减少请求次数；不影响最终输出顺序
    clean_codes = list(dict.fromkeys(clean_codes))

    if not clean_codes:
        return {}

    total = len(clean_codes)
    for position, code in enumerate(clean_codes, start=1):
        high_price = fetch_200_day_high_one(code, use_system_proxy=True)

        if high_price is None:
            logging.warning(f"{code} 使用系统代理获取近200日最高价失败，改为不读取系统代理再试")
            high_price = fetch_200_day_high_one(code, use_system_proxy=False)

        if high_price is not None:
            result[code] = high_price

        progress = 30 + round(45 * position / total)
        logging.info(
            f"进度：{progress}%（近200日高点已处理 {position} 只，共 {total} 只）"
        )

        time.sleep(0.2)

    return result


def format_excel(output_file):
    """
    美化 Excel：
    1. 冻结首行
    2. 自动列宽
    3. 百分比格式
    4. 清仓后涨幅为正标红，为负标绿
    """
    wb = load_workbook(output_file)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_map = {}

    for cell in ws[1]:
        header_map[cell.value] = cell.column
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    for col_name in [AFTER_CLEAR_GAIN_COL, "收益率", HIGH_200_DRAWDOWN_COL]:
        if col_name in header_map:
            col_idx = header_map[col_name]

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = "0.00%"

                if col_name == AFTER_CLEAR_GAIN_COL:
                    try:
                        value = float(cell.value)

                        if value > 0:
                            cell.fill = red_fill
                        elif value < 0:
                            cell.fill = green_fill

                    except Exception:
                        pass

    for col_name in ["成本", CLEAR_PRICE_COL, CURRENT_PRICE_COL, HIGH_200_COL]:
        if col_name in header_map:
            col_idx = header_map[col_name]

            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "0.000"

    wb.save(output_file)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="计算清仓后涨幅与近 200 日高点回撤")
    parser.add_argument("--input", help="清仓分析 Excel；默认使用统一配置")
    parser.add_argument("--output", help="输出 Excel；默认使用带时间戳的文件名")
    parser.add_argument("--log-file", help="轮转日志文件；默认写入统一日志目录")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    input_file = resolve_input(args.input, config_key="clearance_analysis")
    output_file = project_path(args.output or OUTPUT_FILE)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    log_file = project_path(args.log_file) if args.log_file else default_log_dir() / "清仓分析.log"
    configure_root_rotating_logging(log_file)

    logging.info(f"开始读取文件：{input_file}")

    df = pd.read_excel(input_file, sheet_name=SHEET_NAME)

    logging.info(f"读取完成，共 {len(df)} 行")
    logging.info(f"字段列表：{list(df.columns)}")
    logging.info("进度：5%（输入文件读取完成）")

    for col in [CODE_COL, CLEAR_PRICE_COL]:
        if col not in df.columns:
            raise ValueError(f"输入文件缺少必要字段：{col}")

    if CURRENT_PRICE_COL not in df.columns:
        df[CURRENT_PRICE_COL] = None

    if AFTER_CLEAR_GAIN_COL not in df.columns:
        df[AFTER_CLEAR_GAIN_COL] = None

    if HIGH_200_COL not in df.columns:
        df[HIGH_200_COL] = None

    if HIGH_200_DRAWDOWN_COL not in df.columns:
        df[HIGH_200_DRAWDOWN_COL] = None

    # 临时辅助列，最后删除
    df["_标准代码"] = df[CODE_COL].apply(normalize_code)

    # 只取非空代码，避免后面空行也被统计进去
    valid_code_series = df["_标准代码"].dropna()
    valid_code_series = valid_code_series[valid_code_series.astype(str).str.len() == 6]

    codes = valid_code_series.tolist()

    logging.info(f"准备获取现价，股票数量：{len(codes)}")

    price_map = fetch_price_map(codes)

    logging.info(f"现价获取完成，成功获取：{len(price_map)} 只")
    logging.info("进度：30%（现价获取阶段完成）")

    logging.info(f"准备获取近200日最高价，股票数量：{len(codes)}")

    high_200_map = fetch_200_day_high_map(codes)

    logging.info(f"近200日最高价获取完成，成功获取：{len(high_200_map)} 只")
    logging.info("进度：75%（近200日高点获取阶段完成）")

    success_count = 0
    fail_codes = []

    # 重点：按原始 df 顺序逐行回填，不排序
    total_rows = len(df)
    for position, (idx, row) in enumerate(df.iterrows(), start=1):
        progress = 75 + round(20 * position / max(total_rows, 1))
        logging.info(
            f"进度：{progress}%（结果已计算 {position} 行，共 {total_rows} 行）"
        )
        raw_code = row[CODE_COL]
        code = row["_标准代码"]

        if not code or pd.isna(code) or len(str(code)) != 6:
            logging.warning(f"第 {idx + 2} 行股票代码为空，跳过")
            continue

        clear_price = pd.to_numeric(row[CLEAR_PRICE_COL], errors="coerce")

        if pd.isna(clear_price) or clear_price <= 0:
            logging.warning(
                f"{raw_code} 清仓价格异常：{row[CLEAR_PRICE_COL]}，跳过"
            )
            continue

        item = price_map.get(code)

        if not item:
            logging.warning(f"{raw_code} 未获取到当前价格")
            fail_codes.append(str(raw_code))
            continue

        current_price = item["price"]

        high_200_price = high_200_map.get(code)

        after_clear_gain = (current_price - clear_price) / clear_price

        df.at[idx, CURRENT_PRICE_COL] = round(current_price, 3)
        df.at[idx, AFTER_CLEAR_GAIN_COL] = after_clear_gain

        if high_200_price is not None:
            # 只做必要兜底：如果盘中现价已经高于近200日K线最高价，则用现价作为近200日最高价
            # 避免出现“现价相比近200日最高价跌幅”为正数的异常情况
            high_200_price = max(float(high_200_price), float(current_price))

            high_200_drawdown = (current_price - high_200_price) / high_200_price

            df.at[idx, HIGH_200_COL] = round(high_200_price, 3)
            df.at[idx, HIGH_200_DRAWDOWN_COL] = high_200_drawdown
        else:
            logging.warning(f"{raw_code} 未获取到近200日最高价")

        success_count += 1

        logging.info(
            f"{raw_code} {row.get(NAME_COL, '')} | "
            f"清仓价={clear_price} | "
            f"现价={current_price} | "
            f"近200日最高价={high_200_price} | "
            f"清仓后涨幅={after_clear_gain:.2%}"
        )

    df.drop(columns=["_标准代码"], inplace=True)

    df[AFTER_CLEAR_GAIN_COL] = pd.to_numeric(
        df[AFTER_CLEAR_GAIN_COL],
        errors="coerce"
    )

    df[HIGH_200_COL] = pd.to_numeric(
        df[HIGH_200_COL],
        errors="coerce"
    )

    df[HIGH_200_DRAWDOWN_COL] = pd.to_numeric(
        df[HIGH_200_DRAWDOWN_COL],
        errors="coerce"
    )

    logging.info("进度：98%（正在写入并格式化 Excel）")

    try:
        df.to_excel(output_file, index=False)
    except PermissionError:
        backup_file = timestamped_output_path(
            output_file.parent,
            "清仓分析_清仓后涨幅",
            timestamp=datetime.now().strftime("%Y%m%d_%H%M%S_%f"),
        )
        logging.warning(
            f"原输出文件被占用，改用新文件名输出：{backup_file}"
        )
        df.to_excel(backup_file, index=False)
        format_excel(backup_file)

        logging.info("=" * 80)
        logging.info(f"处理完成：{backup_file}")
        logging.info(f"成功计算：{success_count} 只")
        logging.info(f"失败数量：{len(fail_codes)} 只")
        logging.info("进度：100%（处理完成）")

        if fail_codes:
            logging.warning("失败代码如下：")
            logging.warning(",".join(fail_codes))

        logging.info("=" * 80)
        return

    format_excel(output_file)

    logging.info("=" * 80)
    logging.info(f"处理完成：{output_file}")
    logging.info(f"成功计算：{success_count} 只")
    logging.info(f"失败数量：{len(fail_codes)} 只")
    logging.info("进度：100%（处理完成）")

    if fail_codes:
        logging.warning("失败代码如下：")
        logging.warning(",".join(fail_codes))

    logging.info("=" * 80)


if __name__ == "__main__":
    main()
