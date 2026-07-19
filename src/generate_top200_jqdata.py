"""使用本地 JQData SDK 生成项目统一的 top200_stocks 选股明细。"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from pipeline_config import config_value, project_path


TECH_COLUMNS = [
    "当前价格", "200日涨幅", "120日涨幅", "60日涨幅", "20日涨幅", "最近15个交易日涨幅",
    "10日涨幅", "5日涨幅", "量比", "当日涨幅", "近5日波动率", "近20日波动率均值",
    "近200交易日最高价", "近200交易日最低价", "相对本年度偏差(%)",
    "近200个交易日单日涨幅超过8%的次数", "上月交易量相比之前20个月交易量均值的增减",
    "最近15个交易日单日涨幅的交易日个数", "上一个交易日交易量相对之前15日均量(%)",
    "相对20日均线偏差(%)", "相对50日均线偏差(%)", "相对150日均线偏差(%)",
    "相对200日均线偏差(%)", "50日均线相对150日均线偏差(%)",
    "50日均线相对200日均线偏差(%)", "150日均线相对200日均线偏差(%)",
    "200日均线连续上涨的交易日数量", "相对近200日低点偏差(%)",
    "相对近200日高点偏差(%)", "近5年高低点波动率(%)", "距近200日最高值交易日数",
    "距历史最高值交易日数", "RS线连续上涨的交易日数量(对标沪深300)",
    "MACD金叉", "MACD死叉", "KDJ金叉", "KDJ死叉", "BOLL下轨", "连续阳线数",
    "连续阴线数", "相对历史高点跌幅", "相对历史低点反弹",
]

PREFERRED_COLUMNS = [
    "股票代码", "名称", "当前价格", "综合排名", "RS排名", "行业", "概念",
    "机构持股占比_流通口径", "机构占比环比增幅", "年初至今涨幅",
    "200日涨幅", "120日涨幅", "60日涨幅", "20日涨幅", "最近15个交易日涨幅",
    "10日涨幅", "5日涨幅", "60/20/15/10/5日涨幅方差", "量比", "当日涨幅",
    "近5日波动率", "近20日波动率均值", "近200交易日最高价", "近200交易日最低价",
    "相对本年度偏差(%)", "近200个交易日单日涨幅超过8%的次数",
    "上月交易量相比之前20个月交易量均值的增减", "最近15个交易日单日涨幅的交易日个数",
    "上一个交易日交易量相对之前15日均量(%)", "相对20日均线偏差(%)",
    "相对50日均线偏差(%)", "相对150日均线偏差(%)", "相对200日均线偏差(%)",
    "50日均线相对150日均线偏差(%)", "50日均线相对200日均线偏差(%)",
    "150日均线相对200日均线偏差(%)", "200日均线连续上涨的交易日数量",
    "相对近200日低点偏差(%)", "相对近200日高点偏差(%)", "近5年高低点波动率(%)",
    "距近200日最高值交易日数", "距历史最高值交易日数",
    "RS线连续上涨的交易日数量(对标沪深300)", "MACD金叉", "MACD死叉", "KDJ金叉",
    "KDJ死叉", "BOLL下轨", "昨日换手率", "连续阳线数", "连续阴线数",
    "相对历史高点跌幅", "相对历史低点反弹", "PE", "PB", "PEG", "ROE",
    "净利润同比增长率", "毛利率", "总市值(亿)", "流通市值(亿)",
    "最新季度利润增速(%)", "最新财报净利率(%)", "净利率",
    "最新季度EPS同比增幅(%)", "涨幅均值", "趋势得分", "波动得分", "动能得分",
    "估值得分", "成长得分", "综合评分",
]

INSTITUTION_KEYWORDS = (
    "基金", "公募", "私募", "保险", "人寿", "资管", "资产管理", "证券", "券商", "信托",
    "QFII", "主权", "社保", "社会保障", "养老金", "年金", "银行", "理财", "外资",
    "合伙企业", "投资公司", "投资管理", "国际", "Banc", "Bank", "Capital", "Asset",
    "基金管理", "企业年金", "社保基金",
)
EXCLUDE_KEYWORDS = ("自然人", "个人", "職工", "员工", "高管", "博士", "先生", "女士")


def load_jqdata():
    try:
        import jqdatasdk as jq
    except ImportError as exc:
        raise RuntimeError("缺少jqdatasdk，请先执行 pip install -r requirements.txt") from exc
    username = os.getenv("JQDATA_USERNAME", "").strip()
    password = os.getenv("JQDATA_PASSWORD", "").strip()
    if not username or not password:
        raise RuntimeError("请先设置环境变量 JQDATA_USERNAME 和 JQDATA_PASSWORD")
    jq.auth(username, password)
    return jq


def parse_date(value: str | None) -> dt.date:
    if not value:
        return dt.date.today()
    try:
        return dt.datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("date必须使用YYYY-MM-DD格式") from exc


def last_trade_day_on_or_before(jq, value: dt.date) -> dt.date:
    days = jq.get_trade_days(end_date=value, count=1)
    if days is None or len(days) == 0:
        raise RuntimeError(f"无法取得{value}之前的交易日")
    return pd.Timestamp(days[-1]).date()


def previous_year_last_trade_day(jq, value: dt.date) -> dt.date:
    return last_trade_day_on_or_before(jq, dt.date(value.year - 1, 12, 31))


def latest_completed_quarter_statdates(today: dt.date) -> tuple[str, str]:
    if today.month <= 3:
        year, quarter = today.year - 1, 4
    elif today.month <= 6:
        year, quarter = today.year, 1
    elif today.month <= 9:
        year, quarter = today.year, 2
    else:
        year, quarter = today.year, 3
    return f"{year}q{quarter}", f"{year - 1}q{quarter}"


def _safe_ratio(a, b):
    return (a - b) / b if pd.notna(a) and pd.notna(b) and b != 0 else np.nan


def _consecutive(values, predicate) -> int:
    count = 0
    for value in reversed(list(values)):
        if pd.notna(value) and predicate(value):
            count += 1
        else:
            break
    return count


def technical_metrics(history: pd.DataFrame, benchmark: pd.Series, end_date: dt.date) -> dict[str, Any] | None:
    """从单只股票一次历史查询中计算用户原脚本的技术字段。"""
    required = {"close", "high", "low", "volume", "money"}
    if history is None or len(history) < 210 or not required.issubset(history.columns):
        return None
    all_data = history.sort_index().copy()
    data = all_data.tail(300)
    close = pd.to_numeric(data["close"], errors="coerce")
    last = close.iloc[-1]
    if pd.isna(last):
        return None
    result: dict[str, Any] = {"当前价格": float(last)}
    for period in (200, 120, 60, 20, 15, 10, 5):
        name = "最近15个交易日涨幅" if period == 15 else f"{period}日涨幅"
        result[name] = _safe_ratio(last, close.iloc[-period - 1]) if len(close) > period else np.nan
    result["量比"] = data["money"].tail(5).mean() / data["money"].tail(60).mean()
    result["当日涨幅"] = _safe_ratio(last, close.iloc[-2])
    intraday = (data["high"] - data["low"]) / close.shift(1)
    result["近5日波动率"] = intraday.tail(5).mean()
    result["近20日波动率均值"] = intraday.tail(20).mean()
    result["近200交易日最高价"] = data["high"].tail(200).max()
    result["近200交易日最低价"] = data["low"].tail(200).min()
    result["近200个交易日单日涨幅超过8%的次数"] = int((close.pct_change().tail(200) > .08).sum())
    result["最近15个交易日单日涨幅的交易日个数"] = int((close.pct_change().tail(15) > 0).sum())
    result["上一个交易日交易量相对之前15日均量(%)"] = _safe_ratio(
        data["volume"].iloc[-2], data["volume"].iloc[-17:-2].mean()
    )

    monthly = all_data[["volume"]].copy()
    monthly.index = pd.to_datetime(monthly.index)
    monthly_sum = monthly["volume"].resample("ME").sum()
    monthly_sum.index = monthly_sum.index.to_period("M")
    previous_month = pd.Period(end_date, freq="M") - 1
    previous_volume = monthly_sum.get(previous_month, np.nan)
    prior_mean = monthly_sum[monthly_sum.index < previous_month].tail(20).mean()
    result["上月交易量相比之前20个月交易量均值的增减"] = _safe_ratio(previous_volume, prior_mean)

    ema12, ema26 = close.ewm(span=12).mean(), close.ewm(span=26).mean()
    macd = ema12 - ema26
    signal = macd.ewm(span=9).mean()
    result["MACD金叉"] = int(macd.iloc[-2] < signal.iloc[-2] and macd.iloc[-1] > signal.iloc[-1])
    result["MACD死叉"] = int(macd.iloc[-2] > signal.iloc[-2] and macd.iloc[-1] < signal.iloc[-1])
    low9, high9 = close.rolling(9).min(), close.rolling(9).max()
    k = ((close - low9) / (high9 - low9) * 100).ewm(com=2).mean()
    d = k.ewm(com=2).mean()
    result["KDJ金叉"] = int(k.iloc[-2] < d.iloc[-2] and k.iloc[-1] > d.iloc[-1])
    result["KDJ死叉"] = int(k.iloc[-2] > d.iloc[-2] and k.iloc[-1] < d.iloc[-1])

    moving = {period: close.rolling(period).mean() for period in (20, 50, 150, 200)}
    result["BOLL下轨"] = int(last <= moving[20].iloc[-1] - 2 * close.rolling(20).std().iloc[-1])
    for period in (20, 50, 150, 200):
        result[f"相对{period}日均线偏差(%)"] = _safe_ratio(last, moving[period].iloc[-1])
    for short, long in ((50, 150), (50, 200), (150, 200)):
        result[f"{short}日均线相对{long}日均线偏差(%)"] = _safe_ratio(
            moving[short].iloc[-1], moving[long].iloc[-1]
        )
    x = np.arange(10)
    slopes = moving[200].rolling(10).apply(
        lambda values: np.polyfit(x, values, 1)[0] if np.isfinite(values).all() else np.nan,
        raw=True,
    )
    result["200日均线连续上涨的交易日数量"] = _consecutive(slopes.dropna(), lambda value: value > 0)
    low200, high200 = close.tail(200).min(), close.tail(200).max()
    result["相对近200日低点偏差(%)"] = _safe_ratio(last, low200)
    result["相对近200日高点偏差(%)"] = _safe_ratio(last, high200)
    year_close = all_data.loc[pd.to_datetime(all_data.index).year == end_date.year, "close"]
    result["相对本年度偏差(%)"] = _safe_ratio(last, year_close.mean())
    five_year_start = pd.Timestamp(end_date) - pd.Timedelta(days=365 * 5 + 10)
    five_year = all_data[pd.to_datetime(all_data.index) >= five_year_start]["close"]
    result["近5年高低点波动率(%)"] = _safe_ratio(five_year.max(), five_year.min())
    all_close = all_data["close"]
    result["相对历史高点跌幅"] = _safe_ratio(last, all_close.max())
    result["相对历史低点反弹"] = _safe_ratio(last, all_close.min())
    result["距近200日最高值交易日数"] = len(close.tail(200)) - 1 - int(np.nanargmax(close.tail(200).values))
    result["距历史最高值交易日数"] = len(all_close) - 1 - int(np.nanargmax(all_close.values))
    signs = np.sign(close.diff().dropna())
    result["连续阳线数"] = _consecutive(signs, lambda value: value > 0)
    result["连续阴线数"] = _consecutive(signs, lambda value: value < 0)
    aligned = pd.concat([all_close.rename("stock"), benchmark.rename("benchmark")], axis=1).dropna()
    rs_diff = (aligned["stock"] / aligned["benchmark"]).diff().dropna()
    result["RS线连续上涨的交易日数量(对标沪深300)"] = _consecutive(rs_diff, lambda value: value > 0)
    return result


def quantile_score(series: pd.Series, higher_better: bool = True) -> pd.Series:
    rank = pd.to_numeric(series, errors="coerce").rank(pct=True)
    return rank if higher_better else 1 - rank


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n]+', "_", str(name)).strip()
    return cleaned[:80] or "未知"


def highlight_workbook(workbook, sheet_name: str = "Sheet1") -> None:
    sheet = workbook[sheet_name]
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    headers = [cell.value for cell in sheet[1]]
    cross_columns = {"MACD金叉", "KDJ金叉", "BOLL下轨", "MACD死叉", "KDJ死叉"}
    for column, name in enumerate(headers, 1):
        if name in {"股票代码", "名称", "行业", "概念"}:
            continue
        values = [(row, sheet.cell(row, column).value) for row in range(2, sheet.max_row + 1)]
        numeric = [(row, value) for row, value in values if isinstance(value, (int, float))]
        if name in cross_columns:
            for row, value in numeric:
                if value == 1:
                    sheet.cell(row, column).fill = green if "死叉" in name else red
        elif len(numeric) >= 20:
            ordered = sorted(numeric, key=lambda item: item[1], reverse=name != "综合排名")
            for row, _ in ordered[:10]:
                sheet.cell(row, column).fill = red
            for row, _ in ordered[-10:]:
                sheet.cell(row, column).fill = green


def dataframe_to_excel_bytes(frame: pd.DataFrame) -> bytes:
    source = BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Sheet1")
    workbook = load_workbook(BytesIO(source.getvalue()))
    highlight_workbook(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _chunked(values: list[str], size: int = 800):
    for start in range(0, len(values), size):
        yield values[start:start + size]


def fetch_fundamentals(jq, codes: list[str], date: dt.date) -> pd.DataFrame:
    frames = []
    for batch in _chunked(codes):
        query = jq.query(
            jq.valuation.code, jq.valuation.pe_ratio, jq.valuation.pb_ratio,
            jq.valuation.market_cap, jq.valuation.circulating_market_cap,
            jq.valuation.turnover_ratio, jq.indicator.roe,
            jq.indicator.inc_net_profit_year_on_year, jq.indicator.gross_profit_margin,
        ).filter(jq.valuation.code.in_(batch))
        frames.append(jq.get_fundamentals(query, date=date))
    frame = pd.concat([item for item in frames if item is not None], ignore_index=True)
    return frame.rename(columns={
        "code": "股票代码", "pe_ratio": "PE", "pb_ratio": "PB",
        "market_cap": "总市值(亿)", "circulating_market_cap": "流通市值(亿)",
        "turnover_ratio": "昨日换手率", "roe": "ROE",
        "inc_net_profit_year_on_year": "净利润同比增长率", "gross_profit_margin": "毛利率",
    })


def fetch_turnover(jq, codes: list[str], date: dt.date) -> dict[str, float]:
    frames = []
    for batch in _chunked(codes):
        query = jq.query(jq.valuation.code, jq.valuation.turnover_ratio).filter(
            jq.valuation.code.in_(batch)
        )
        frames.append(jq.get_fundamentals(query, date=date))
    frame = pd.concat([item for item in frames if item is not None], ignore_index=True)
    return frame.set_index("code")["turnover_ratio"].to_dict() if not frame.empty else {}


def fetch_quarter_maps(jq, codes: list[str], today: dt.date) -> tuple[dict, dict, dict]:
    current, previous = latest_completed_quarter_statdates(today)

    def fetch(stat_date: str) -> pd.DataFrame:
        frames = []
        for batch in _chunked(codes):
            query = jq.query(
                jq.income.code, jq.income.net_profit, jq.income.operating_revenue,
                jq.income.basic_eps,
            ).filter(jq.income.code.in_(batch))
            frames.append(jq.get_fundamentals(query, statDate=stat_date))
        return pd.concat([item for item in frames if item is not None], ignore_index=True)

    current_frame, previous_frame = fetch(current), fetch(previous)
    merged = current_frame.merge(previous_frame, on="code", how="outer", suffixes=("_cur", "_prev"))
    profit = merged.set_index("code").apply(
        lambda row: _safe_ratio(row.get("net_profit_cur"), row.get("net_profit_prev")), axis=1
    ).to_dict()
    eps = merged.set_index("code").apply(
        lambda row: _safe_ratio(row.get("basic_eps_cur"), row.get("basic_eps_prev")), axis=1
    ).to_dict()
    margin = current_frame.set_index("code").apply(
        lambda row: row.get("net_profit") / row.get("operating_revenue")
        if pd.notna(row.get("operating_revenue")) and row.get("operating_revenue") != 0 else np.nan,
        axis=1,
    ).to_dict()
    return profit, margin, eps


def build_industry_and_concepts(jq, date: dt.date) -> tuple[dict[str, str], dict[str, str]]:
    industry_map: dict[str, str] = {}
    industries = jq.get_industries(name="sw_l1")
    for code, row in industries.iterrows():
        for stock in jq.get_industry_stocks(code, date=date):
            industry_map[stock] = row.get("name", str(code))
    concept_map: dict[str, list[str]] = {}
    concepts = jq.get_concepts()
    for code, row in concepts.iterrows():
        try:
            stocks = jq.get_concept_stocks(code, date=date)
        except Exception:
            continue
        for stock in stocks:
            concept_map.setdefault(stock, []).append(row.get("name", str(code)))
    return industry_map, {
        stock: "、".join(sorted(set(names))) for stock, names in concept_map.items()
    }


def _is_institution(name: str) -> bool:
    return bool(name) and not any(key in name for key in EXCLUDE_KEYWORDS) and any(
        key in name for key in INSTITUTION_KEYWORDS
    )


def _parse_number(value) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).replace(",", "").strip()
    multiplier = 1
    if text.endswith("亿"):
        text, multiplier = text[:-1], 1e8
    elif text.endswith("万"):
        text, multiplier = text[:-1], 1e4
    try:
        return float(text.rstrip("%")) * multiplier / (100 if text.endswith("%") else 1)
    except ValueError:
        return None


def public_fund_ratio(jq, code: str, end_date: dt.date) -> float | None:
    try:
        table = jq.finance.FUND_PORTFOLIO_STOCK
        latest = jq.finance.run_query(
            jq.query(table.period).filter(table.code == code, table.period <= end_date)
            .order_by(table.period.desc()).limit(1)
        )
        if latest is None or latest.empty:
            return np.nan
        period = pd.Timestamp(latest.iloc[0]["period"]).date()
        holdings = jq.finance.run_query(
            jq.query(table).filter(table.code == code, table.period == period)
        )
        denominator = jq.get_fundamentals(
            jq.query(jq.valuation.circulating_cap).filter(jq.valuation.code == code),
            date=last_trade_day_on_or_before(jq, period),
        )
        shares = denominator.iloc[0].get("circulating_cap", np.nan) * 1e8 if not denominator.empty else np.nan
        if pd.isna(shares) or not shares:
            return np.nan
        share_columns = [column for column in holdings if any(
            key in column.lower() for key in ("share_number", "shares", "volume", "hold_amount", "持股", "持仓")
        )]
        total = 0.0
        for _, row in holdings.iterrows():
            held = next((_parse_number(row.get(column)) for column in share_columns if _parse_number(row.get(column)) is not None), None)
            total += held or 0
        return float(np.clip(total / shares, 0, 1))
    except Exception:
        return np.nan


def institutional_ratios(jq, code: str, end_date: dt.date) -> tuple[float | None, float | None]:
    """保留原脚本的流通前十大机构口径；无权限时返回空值。"""
    try:
        table = jq.finance.STK_SHAREHOLDER_FLOATING_TOP10
        dates = jq.finance.run_query(
            jq.query(table.report_date).filter(table.code == code, table.report_date <= end_date)
            .order_by(table.report_date.desc()).limit(2)
        )
        report_dates = sorted(pd.to_datetime(dates["report_date"]).dt.date.unique(), reverse=True)
        ratios = []
        for report_date in report_dates[:2]:
            holders = jq.finance.run_query(
                jq.query(table).filter(table.code == code, table.report_date == report_date)
            )
            denominator = jq.get_fundamentals(
                jq.query(jq.valuation.circulating_cap).filter(jq.valuation.code == code),
                date=last_trade_day_on_or_before(jq, report_date),
            )
            shares = denominator.iloc[0].get("circulating_cap", np.nan) * 1e8 if not denominator.empty else np.nan
            name_columns = [column for column in holders if any(key in column.lower() for key in ("holder", "shareholder", "name", "股东", "名称"))]
            share_columns = [column for column in holders if any(key in column.lower() for key in ("share_number", "shares", "hold_amount", "持股"))]
            ratio_columns = [column for column in holders if any(key in column.lower() for key in ("ratio", "percent", "proportion", "占比", "比例"))]
            total = 0.0
            for _, row in holders.iterrows():
                name = str(row.get(name_columns[0], "")) if name_columns else ""
                if not _is_institution(name):
                    continue
                held = next((_parse_number(row.get(column)) for column in share_columns if _parse_number(row.get(column)) is not None), None)
                ratio = next((_parse_number(row.get(column)) for column in ratio_columns if _parse_number(row.get(column)) is not None), None)
                total += held if held is not None else ((ratio or 0) * shares if pd.notna(shares) else 0)
            ratios.append(float(np.clip(total / shares, 0, 1)) if pd.notna(shares) and shares else np.nan)
        current = ratios[0] if ratios else np.nan
        previous = ratios[1] if len(ratios) > 1 else np.nan
        if pd.isna(current):
            current = public_fund_ratio(jq, code, end_date)
        return current, current - previous if pd.notna(current) and pd.notna(previous) else np.nan
    except Exception:
        return public_fund_ratio(jq, code, end_date), np.nan


def score_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    return_columns = ["年初至今涨幅", "200日涨幅", "120日涨幅", "60日涨幅", "20日涨幅", "10日涨幅", "5日涨幅"]
    result["涨幅均值"] = result[return_columns].apply(pd.to_numeric, errors="coerce").mean(axis=1)
    result["趋势得分"] = quantile_score(result["涨幅均值"])
    result["波动得分"] = quantile_score(result["近5日波动率"], False)
    result["动能得分"] = quantile_score(result["量比"])
    result["估值得分"] = (
        quantile_score(result["PE"], False) + quantile_score(result["PB"], False)
        + quantile_score(result["PEG"], False)
    ) / 3
    result["成长得分"] = (quantile_score(result["ROE"]) + quantile_score(result["净利润同比增长率"])) / 2
    result["综合评分"] = (
        result["趋势得分"] * .25 + result["波动得分"] * .10 + result["动能得分"] * .15
        + result["估值得分"] * .20 + result["成长得分"] * .30
    )
    result["综合排名"] = result["综合评分"].rank(ascending=False, method="min")
    result["RS排名"] = result["涨幅均值"].rank(pct=True).mul(100).round().astype("Int64")
    return result


def generate(jq, *, requested_date: dt.date, limit: int, include_institutions: bool) -> tuple[pd.DataFrame, dt.date, dt.date]:
    end_date = last_trade_day_on_or_before(jq, requested_date)
    base_date = previous_year_last_trade_day(jq, end_date)
    securities = jq.get_all_securities(["stock"], date=end_date)
    names = securities["display_name"].astype(str)
    securities = securities[~names.str.contains("ST", case=False, na=False)]
    securities = securities[~securities.index.str.startswith("688")]
    securities = securities[pd.to_datetime(securities["end_date"]).dt.date >= end_date]
    benchmark_frame = jq.get_price(
        "000300.XSHG", start_date="2015-01-01", end_date=end_date,
        frequency="daily", fields=["close"], fq="none",
    )
    benchmark = benchmark_frame["close"]
    rows = []
    total = len(securities)
    for position, (code, security) in enumerate(securities.iterrows(), 1):
        try:
            history = jq.get_price(
                code, start_date="2015-01-01", end_date=end_date, frequency="daily",
                fields=["close", "money", "volume", "high", "low"], fq="none",
            )
            metrics = technical_metrics(history, benchmark, end_date)
            if metrics is None:
                continue
            before = history[pd.to_datetime(history.index).date <= base_date]
            if before.empty:
                continue
            metrics.update({
                "股票代码": code, "名称": security["display_name"],
                "年初至今涨幅": _safe_ratio(metrics["当前价格"], before["close"].iloc[-1]),
            })
            rows.append(metrics)
        except Exception as exc:
            print(f"[跳过] {code} {security['display_name']}：{exc}")
        if position % 25 == 0 or position == total:
            print(f"已处理 {position}/{total} 只股票")
    if not rows:
        raise RuntimeError("没有取得可用股票数据，请检查账号权限和查询日期")
    frame = pd.DataFrame(rows).sort_values("年初至今涨幅", ascending=False).head(limit).reset_index(drop=True)
    codes = frame["股票代码"].tolist()
    fundamentals = fetch_fundamentals(jq, codes, end_date)
    previous_trade_days = jq.get_trade_days(end_date=end_date, count=2)
    previous_trade_day = pd.Timestamp(previous_trade_days[0]).date()
    fundamentals["昨日换手率"] = fundamentals["股票代码"].map(
        fetch_turnover(jq, codes, previous_trade_day)
    )
    fundamentals["流通市值(亿)"] = fundamentals["流通市值(亿)"].fillna(fundamentals["总市值(亿)"])
    fundamentals["PEG"] = fundamentals["PE"] / fundamentals["净利润同比增长率"].replace(0, np.nan)
    industry, concepts = build_industry_and_concepts(jq, end_date)
    fundamentals["行业"] = fundamentals["股票代码"].map(industry).fillna("未知行业")
    fundamentals["概念"] = fundamentals["股票代码"].map(concepts)
    frame = frame.merge(fundamentals, on="股票代码", how="left")
    profit, margin, eps = fetch_quarter_maps(jq, codes, end_date)
    frame["最新季度利润增速(%)"] = frame["股票代码"].map(profit)
    frame["最新财报净利率(%)"] = frame["股票代码"].map(margin)
    frame["最新季度EPS同比增幅(%)"] = frame["股票代码"].map(eps)
    frame["净利率"] = frame["最新财报净利率(%)"]
    returns = ["60日涨幅", "20日涨幅", "最近15个交易日涨幅", "10日涨幅", "5日涨幅"]
    frame["60/20/15/10/5日涨幅方差"] = frame[returns].apply(pd.to_numeric, errors="coerce").var(axis=1, ddof=0)
    if include_institutions:
        values = []
        for index, code in enumerate(codes, 1):
            values.append(institutional_ratios(jq, code, end_date))
            if index % 25 == 0 or index == len(codes):
                print(f"机构数据 {index}/{len(codes)}")
        frame["机构持股占比_流通口径"] = [item[0] for item in values]
        frame["机构占比环比增幅"] = [item[1] for item in values]
    else:
        frame["机构持股占比_流通口径"] = np.nan
        frame["机构占比环比增幅"] = np.nan
    frame = score_frame(frame)
    for column in PREFERRED_COLUMNS:
        if column not in frame:
            frame[column] = np.nan
    frame = frame[PREFERRED_COLUMNS]
    float_columns = frame.select_dtypes(include=["float"]).columns
    frame[float_columns] = frame[float_columns].round(3)
    return frame, end_date, base_date


def save_outputs(frame: pd.DataFrame, data_date: dt.date, input_dir: Path, output_dir: Path) -> tuple[Path, Path]:
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    tag = data_date.strftime("%Y%m%d")
    excel_path = input_dir / f"top200_stocks_{tag}.xlsx"
    zip_path = output_dir / f"top200_stocks_{tag}_summary_and_industries.zip"
    excel_bytes = dataframe_to_excel_bytes(frame)
    temporary = excel_path.with_suffix(".xlsx.tmp")
    temporary.write_bytes(excel_bytes)
    temporary.replace(excel_path)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(excel_path.name, excel_bytes)
        grouped = frame.copy()
        grouped["行业"] = grouped["行业"].fillna("未知行业")
        for industry, subset in grouped.groupby("行业"):
            archive.writestr(
                f"top200_stocks_{tag}_{sanitize_filename(industry)}.xlsx",
                dataframe_to_excel_bytes(subset),
            )
    return excel_path, zip_path


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="通过JQData生成Top200选股明细和行业ZIP")
    parser.add_argument("--date", help="目标日期YYYY-MM-DD；默认今天并向前对齐交易日")
    parser.add_argument("--limit", type=int, default=5000, help="最多保留股票数；原逻辑为5000")
    parser.add_argument("--skip-institutions", action="store_true", help="跳过耗时且可能受权限限制的机构持股查询")
    parser.add_argument("--input-dir", default=config_value("files", "input_dir", "data/input"))
    parser.add_argument("--output-dir", default=config_value("files", "output_dir", "data/output"))
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    if args.limit <= 0:
        raise ValueError("limit必须为正数")
    jq = load_jqdata()
    frame, data_date, base_date = generate(
        jq, requested_date=parse_date(args.date), limit=args.limit,
        include_institutions=not args.skip_institutions,
    )
    excel_path, zip_path = save_outputs(
        frame, data_date, project_path(args.input_dir), project_path(args.output_dir)
    )
    print(f"选股明细：{excel_path}")
    print(f"行业压缩包：{zip_path}")
    print(f"股票数量：{len(frame)} | 数据交易日：{data_date} | 年初基准日：{base_date}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
