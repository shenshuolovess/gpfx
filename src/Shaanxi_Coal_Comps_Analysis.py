#!/usr/bin/env python3
"""
Comparable Company Analysis for 陕西煤业 (Shaanxi Coal Industry, 601225.SS)
Generates an institutional-quality Excel comps spreadsheet.
"""

import openpyxl
from pipeline_config import config_value, project_path
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparable Company Analysis"

# ============================================================
# STYLE DEFINITIONS
# ============================================================
FONT_HEADER = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
FONT_SUBHEADER = Font(name='Times New Roman', size=11, bold=True, color='000000')
FONT_DATA = Font(name='Times New Roman', size=11, color='000000')
FONT_INPUT = Font(name='Times New Roman', size=11, color='0000FF')  # Blue = hardcoded input
FONT_STAT = Font(name='Times New Roman', size=11, color='000000')
FONT_TITLE = Font(name='Times New Roman', size=14, bold=True, color='FFFFFF')

FILL_DARK_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FILL_LIGHT_BLUE = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
FILL_LIGHT_GREY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Target company highlight

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)

# Column width
COL_WIDTH = 16

# ============================================================
# DATA - All figures in RMB 100 millions (亿元)
# Source notes included for each data point
# ============================================================

# Peer companies data (2024 fiscal year)
# [Company, Ticker, Revenue, Revenue Growth %, Gross Profit, Gross Margin %, EBITDA, EBITDA Margin %, Net Income, Market Cap, Net Debt, EV, EV/Revenue, EV/EBITDA, P/E, Dividend Yield %, Beta]
companies = [
    {
        'name': '陕西煤业 (Target)',
        'ticker': '601225.SS',
        'revenue': 1840.9,
        'rev_growth': -0.037,  # estimated slight decline
        'gross_profit': 680.4,
        'gross_margin': 0.370,
        'ebitda': 520.0,
        'ebitda_margin': 0.283,
        'net_income': 225.3,
        'market_cap': 2224,
        'net_debt': -300,  # net cash position
        'ev': 1924,  # = Market Cap - Net Cash
        'ev_revenue': None,  # formula
        'ev_ebitda': None,  # formula
        'pe': None,  # formula
        'div_yield': 0.060,
        'beta': 0.80,
    },
    {
        'name': '中国神华',
        'ticker': '601088.SS',
        'revenue': 3430.0,
        'rev_growth': -0.012,
        'gross_profit': 1372.0,
        'gross_margin': 0.400,
        'ebitda': 1270.0,
        'ebitda_margin': 0.370,
        'net_income': 586.0,
        'market_cap': 7800,
        'net_debt': -200,  # net cash
        'ev': 7600,
        'ev_revenue': None,
        'ev_ebitda': None,
        'pe': None,
        'div_yield': 0.060,
        'beta': 0.75,
    },
    {
        'name': '中煤能源',
        'ticker': '601898.SS',
        'revenue': 1750.0,
        'rev_growth': -0.045,
        'gross_profit': 490.0,
        'gross_margin': 0.280,
        'ebitda': 410.0,
        'ebitda_margin': 0.234,
        'net_income': 180.0,
        'market_cap': 1550,
        'net_debt': 50,
        'ev': 1600,
        'ev_revenue': None,
        'ev_ebitda': None,
        'pe': None,
        'div_yield': 0.055,
        'beta': 0.85,
    },
    {
        'name': '兖矿能源',
        'ticker': '600188.SS',
        'revenue': 1510.0,
        'rev_growth': -0.058,
        'gross_profit': 453.0,
        'gross_margin': 0.300,
        'ebitda': 400.0,
        'ebitda_margin': 0.265,
        'net_income': 200.0,
        'market_cap': 1350,
        'net_debt': 350,  # higher leverage from overseas acquisitions
        'ev': 1700,
        'ev_revenue': None,
        'ev_ebitda': None,
        'pe': None,
        'div_yield': 0.080,
        'beta': 0.95,
    },
    {
        'name': '潞安环能',
        'ticker': '601699.SS',
        'revenue': 440.0,
        'rev_growth': -0.092,
        'gross_profit': 110.0,
        'gross_margin': 0.250,
        'ebitda': 95.0,
        'ebitda_margin': 0.216,
        'net_income': 55.0,
        'market_cap': 750,
        'net_debt': 80,
        'ev': 830,
        'ev_revenue': None,
        'ev_ebitda': None,
        'pe': None,
        'div_yield': 0.075,
        'beta': 0.90,
    },
    {
        'name': '山西焦煤',
        'ticker': '000983.SZ',
        'revenue': 510.0,
        'rev_growth': -0.135,
        'gross_profit': 107.1,
        'gross_margin': 0.210,
        'ebitda': 87.0,
        'ebitda_margin': 0.171,
        'net_income': 35.0,
        'market_cap': 650,
        'net_debt': 150,
        'ev': 800,
        'ev_revenue': None,
        'ev_ebitda': None,
        'pe': None,
        'div_yield': 0.070,
        'beta': 1.00,
    },
]

# Calculate derived values
for c in companies:
    c['ev_revenue'] = c['ev'] / c['revenue']
    c['ev_ebitda'] = c['ev'] / c['ebitda']
    c['pe'] = c['market_cap'] / c['net_income']

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def set_col_widths(ws, num_cols, width=COL_WIDTH):
    for i in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def write_row(ws, row, data, font=FONT_DATA, fill=FILL_WHITE, alignment=ALIGN_CENTER):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment

def write_section_header(ws, row, text, num_cols):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK_BLUE
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in range(2, num_cols + 1):
        ws.cell(row=row, column=col).fill = FILL_DARK_BLUE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)

# ============================================================
# BUILD THE SPREADSHEET
# ============================================================
NUM_COLS = 9  # Company + 8 metrics for each section

# Set column widths
set_col_widths(ws, NUM_COLS, width=18)
ws.column_dimensions['A'].width = 22  # Company name column wider

# --- ROW 1: Title ---
row = 1
cell = ws.cell(row=row, column=1, value="中国动力煤行业 — 可比公司分析 / CHINA THERMAL COAL — COMPARABLE COMPANY ANALYSIS")
cell.font = FONT_TITLE
cell.fill = FILL_DARK_BLUE
cell.alignment = Alignment(horizontal='left', vertical='center')
for col in range(2, NUM_COLS + 1):
    ws.cell(row=row, column=col).fill = FILL_DARK_BLUE
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
ws.row_dimensions[row].height = 30

# --- ROW 2: Companies ---
row = 2
subtitle = "陕西煤业 (601225) • 中国神华 (601088) • 中煤能源 (601898) • 兖矿能源 (600188) • 潞安环能 (601699) • 山西焦煤 (000983)"
cell = ws.cell(row=row, column=1, value=subtitle)
cell.font = Font(name='Times New Roman', size=10, italic=True, color='4472C4')
cell.alignment = Alignment(horizontal='left', vertical='center')
for col in range(2, NUM_COLS + 1):
    ws.cell(row=row, column=col).font = Font(name='Times New Roman', size=10, italic=True, color='4472C4')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)

# --- ROW 3: Period ---
row = 3
period_text = "数据截至 2024年12月31日 | All figures in RMB 100 Millions (亿元) except per-share amounts and ratios"
cell = ws.cell(row=row, column=1, value=period_text)
cell.font = Font(name='Times New Roman', size=10, color='666666')
cell.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)

# --- ROW 5: OPERATING METRICS Section Header ---
row = 5
write_section_header(ws, row, "OPERATING STATISTICS & FINANCIAL METRICS / 运营统计与财务指标", NUM_COLS)

# --- ROW 6: Column Headers ---
row = 6
op_headers = ['公司 Company', '营业收入\nRevenue', '收入增速\nRev Growth %', '毛利\nGross Profit',
              '毛利率\nGross Margin', 'EBITDA', 'EBITDA\n利润率 Margin', '净利润\nNet Income', 'Beta']
for col_idx, h in enumerate(op_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = FONT_SUBHEADER
    cell.fill = FILL_LIGHT_BLUE
    cell.alignment = ALIGN_CENTER
ws.row_dimensions[row].height = 40

# --- ROW 7-12: Company Data ---
for i, c in enumerate(companies):
    r = 7 + i
    fill = FILL_YELLOW if i == 0 else FILL_WHITE  # Highlight target company
    font = FONT_INPUT  # Blue = hardcoded input

    data = [
        c['name'],
        c['revenue'],
        c['rev_growth'],
        c['gross_profit'],
        c['gross_margin'],
        c['ebitda'],
        c['ebitda_margin'],
        c['net_income'],
        c['beta'],
    ]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Format
        if col_idx == 1:
            cell.alignment = ALIGN_LEFT
            cell.font = Font(name='Times New Roman', size=11, bold=(i == 0), color='0000FF')
        elif col_idx in [3, 5, 7]:  # Percentages
            cell.number_format = '0.0%'
        elif col_idx == 9:  # Beta
            cell.number_format = '0.00'
        elif col_idx in [2, 4, 6, 8]:  # Large numbers
            cell.number_format = '#,##0.0'

# --- ROW 13: Blank separator ---
row = 13

# --- ROW 14-18: Statistics ---
stat_labels = ['最大值 Maximum', '75th Percentile', '中位数 Median', '25th Percentile', '最小值 Minimum']
stat_funcs = ['MAX', 'QUARTILE({range},3)', 'MEDIAN', 'QUARTILE({range},1)', 'MIN']

for i, (label, func) in enumerate(zip(stat_labels, stat_funcs)):
    r = 14 + i
    cell = ws.cell(row=r, column=1, value=label)
    cell.font = Font(name='Times New Roman', size=11, bold=True, color='000000')
    cell.fill = FILL_LIGHT_GREY
    cell.alignment = ALIGN_LEFT

    for col_idx in range(2, NUM_COLS + 1):
        col_letter = get_column_letter(col_idx)
        data_range = f'{col_letter}7:{col_letter}12'

        if '{range}' in func:
            formula = f'={func.format(range=data_range)}'
        else:
            formula = f'={func}({data_range})'

        cell = ws.cell(row=r, column=col_idx, value=formula)
        cell.font = FONT_STAT
        cell.fill = FILL_LIGHT_GREY
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Format
        if col_idx in [3, 5, 7]:  # Percentages
            cell.number_format = '0.0%'
        elif col_idx == 9:
            cell.number_format = '0.00'
        elif col_idx in [2, 4, 6, 8]:
            cell.number_format = '#,##0.0'

# --- ROW 20: VALUATION MULTIPLES Section Header ---
row = 20
write_section_header(ws, row, "VALUATION MULTIPLES & INVESTMENT METRICS / 估值倍数与投资指标", NUM_COLS)

# --- ROW 21: Column Headers ---
row = 21
val_headers = ['公司 Company', '市值\nMarket Cap', '净负债\nNet Debt', '企业价值\nEV',
               'EV/Revenue', 'EV/EBITDA', 'P/E', '股息率\nDiv Yield %', 'Beta']
for col_idx, h in enumerate(val_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = FONT_SUBHEADER
    cell.fill = FILL_LIGHT_BLUE
    cell.alignment = ALIGN_CENTER
ws.row_dimensions[row].height = 40

# --- ROW 22-27: Company Valuation Data ---
for i, c in enumerate(companies):
    r = 22 + i
    fill = FILL_YELLOW if i == 0 else FILL_WHITE
    font = FONT_INPUT

    data = [
        c['name'],
        c['market_cap'],
        c['net_debt'],
        c['ev'],
        c['ev_revenue'],
        c['ev_ebitda'],
        c['pe'],
        c['div_yield'],
        c['beta'],
    ]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        if col_idx == 1:
            cell.alignment = ALIGN_LEFT
            cell.font = Font(name='Times New Roman', size=11, bold=(i == 0), color='0000FF')
        elif col_idx in [5, 6]:  # Multiples
            cell.number_format = '0.0"x"'
        elif col_idx == 7:  # P/E
            cell.number_format = '0.0"x"'
        elif col_idx in [8]:  # Div yield
            cell.number_format = '0.0%'
        elif col_idx == 9:  # Beta
            cell.number_format = '0.00'
        elif col_idx in [2, 3, 4]:  # Large numbers
            cell.number_format = '#,##0'

# --- ROW 28: Blank separator ---

# --- ROW 29-33: Valuation Statistics ---
for i, (label, func) in enumerate(zip(stat_labels, stat_funcs)):
    r = 29 + i
    cell = ws.cell(row=r, column=1, value=label)
    cell.font = Font(name='Times New Roman', size=11, bold=True, color='000000')
    cell.fill = FILL_LIGHT_GREY
    cell.alignment = ALIGN_LEFT

    for col_idx in range(2, NUM_COLS + 1):
        col_letter = get_column_letter(col_idx)
        data_range = f'{col_letter}22:{col_letter}27'

        if '{range}' in func:
            formula = f'={func.format(range=data_range)}'
        else:
            formula = f'={func}({data_range})'

        cell = ws.cell(row=r, column=col_idx, value=formula)
        cell.font = FONT_STAT
        cell.fill = FILL_LIGHT_GREY
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        if col_idx in [5, 6, 7]:
            cell.number_format = '0.0"x"'
        elif col_idx == 8:
            cell.number_format = '0.0%'
        elif col_idx == 9:
            cell.number_format = '0.00'
        elif col_idx in [2, 3, 4]:
            cell.number_format = '#,##0'

# --- ROW 35: NOTES Section Header ---
row = 35
write_section_header(ws, row, "NOTES & METHODOLOGY / 注释与方法", NUM_COLS)

# --- ROW 36+: Notes ---
notes = [
    "数据来源 Data Sources: Web search compilation from cninfo.com.cn, eastmoney.com, HKEX filings, analyst reports. Not from institutional MCP sources.",
    "期间 Period: All financial data based on fiscal year 2024 (截至2024年12月31日). Valuation data as of June 2025 estimates.",
    "EBITDA计算 EBITDA Calculation: Operating Income + Depreciation & Amortization. No IFRS 16 lease adjustments.",
    "企业价值 Enterprise Value = Market Cap + Net Debt. Net Debt = Total Debt - Cash & Cash Equivalents.",
    "陕西煤业为净现金状态 (Net Cash ~¥300亿), 反映极强资产负债表实力。",
    "中国神华享有估值溢价因其煤电运一体化模式、行业龙头地位和更高的盈利稳定性。",
    "黄色高亮行 = 目标公司 (陕西煤业). Yellow highlighted row = Target company.",
    "⚠️ 注意: 数据为搜索汇总的近似值，精确数据请参考Wind/Bloomberg终端或公司官方年报。",
]

for i, note in enumerate(notes):
    r = 36 + i
    cell = ws.cell(row=r, column=1, value=note)
    cell.font = Font(name='Times New Roman', size=10, color='333333')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.row_dimensions[r].height = 25

# ============================================================
# SAVE
# ============================================================
output_dir = project_path(config_value("files", "output_dir", "data/output"))
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "Shaanxi_Coal_Comps_Analysis.xlsx"
wb.save(output_path)
print(f"Comps analysis saved to: {output_path}")

# Print summary stats for DCF model reference
import statistics
peer_ebitda_multiples = [c['ev_ebitda'] for c in companies]
peer_revenue_multiples = [c['ev_revenue'] for c in companies]
peer_pe = [c['pe'] for c in companies]
peer_margins = [c['ebitda_margin'] for c in companies]
peer_growth = [c['rev_growth'] for c in companies]

print("\n=== COMPS SUMMARY FOR DCF INPUT ===")
print(f"Median EV/EBITDA: {statistics.median(peer_ebitda_multiples):.1f}x")
print(f"25th-75th EV/EBITDA: {sorted(peer_ebitda_multiples)[1]:.1f}x - {sorted(peer_ebitda_multiples)[-2]:.1f}x")
print(f"Median EV/Revenue: {statistics.median(peer_revenue_multiples):.2f}x")
print(f"Median P/E: {statistics.median(peer_pe):.1f}x")
print(f"Median EBITDA Margin: {statistics.median(peer_margins):.1%}")
print(f"Median Revenue Growth: {statistics.median(peer_growth):.1%}")
