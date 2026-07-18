import openpyxl
from pipeline_config import config_value, project_path
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Semiconductor Comps"

# ============================================================
# STYLE DEFINITIONS
# ============================================================
font_tnr = "Times New Roman"
dark_blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
light_blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_title = Font(name=font_tnr, size=14, bold=True, color="FFFFFF")
font_section = Font(name=font_tnr, size=12, bold=True, color="FFFFFF")
font_col_header = Font(name=font_tnr, size=11, bold=True, color="000000")
font_stat_label = Font(name=font_tnr, size=11, bold=True, color="000000")
font_company = Font(name=font_tnr, size=11, bold=True, color="000000")
font_data = Font(name=font_tnr, size=11, color="000000")
font_input = Font(name=font_tnr, size=11, color="0000B0")

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")

col_widths = {
    'A': 22, 'B': 18, 'C': 16, 'D': 16, 'E': 14,
    'F': 16, 'G': 14, 'H': 16, 'I': 14,
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ============================================================
# ROW 1-3: HEADER BLOCK
# ============================================================
ws.merge_cells("A1:I1")
ws["A1"] = "SEMICONDUCTOR — COMPARABLE COMPANY ANALYSIS"
ws["A1"].font = font_title
ws["A1"].fill = dark_blue_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:I2")
ws["A2"] = "Advanced Micro Devices (AMD) • NVIDIA (NVDA) • Broadcom (AVGO) • Qualcomm (QCOM) • Intel (INTC) • Micron (MU)"
ws["A2"].font = Font(name=font_tnr, size=11, italic=True, color="FFFFFF")
ws["A2"].fill = dark_blue_fill
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:I3")
ws["A3"] = "As of June 25, 2026 | All figures in USD Millions except per-share amounts and ratios"
ws["A3"].font = Font(name=font_tnr, size=10, color="FFFFFF")
ws["A3"].fill = dark_blue_fill
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

for row in range(1, 4):
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = dark_blue_fill

# ============================================================
# ROW 5: SECTION HEADER — OPERATING STATISTICS
# ============================================================
row = 5
ws.merge_cells(f"A{row}:I{row}")
ws[f"A{row}"] = "OPERATING STATISTICS & FINANCIAL METRICS"
ws[f"A{row}"].font = font_section
ws[f"A{row}"].fill = dark_blue_fill
ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 10):
    ws.cell(row=row, column=col).fill = dark_blue_fill

# ============================================================
# ROW 6: COLUMN HEADERS
# ============================================================
row = 6
headers = [
    "Company", "Revenue\n(LTM)", "Revenue\nGrowth (YoY)", "Gross Profit\n(LTM)",
    "Gross\nMargin", "EBITDA\n(LTM)", "EBITDA\nMargin", "Net Income\n(LTM)", "EPS\n(Diluted)"
]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = font_col_header
    cell.fill = light_blue_fill
    cell.alignment = center_align
ws.row_dimensions[row].height = 35

# ============================================================
# RAW DATA ROWS (7-12)
# ============================================================
companies = [
    {
        "name": "AMD", "ticker": "AMD",
        "revenue": 37454, "rev_growth": 0.350,
        "gross_margin": 0.5306, "ebitda": 7430, "ebitda_margin": 0.1984,
        "net_income": 5010, "eps": 3.05,
        "comment": "Source: AMD Q1 FY2026 10-Q filing, SEC EDGAR, accessed 2026-06-25. TTM = Q2-Q4 FY2025 + Q1 FY2026. https://ir.amd.com",
    },
    {
        "name": "NVIDIA", "ticker": "NVDA",
        "revenue": 253490, "rev_growth": 0.852,
        "gross_margin": 0.7415, "ebitda": 165510, "ebitda_margin": 0.6530,
        "net_income": 159610, "eps": 6.53,
        "comment": "Source: NVIDIA FY2026 10-K filing, SEC EDGAR, accessed 2026-06-25. https://investor.nvidia.com",
    },
    {
        "name": "Broadcom", "ticker": "AVGO",
        "revenue": 75470, "rev_growth": 0.480,
        "gross_margin": 0.7628, "ebitda": 42080, "ebitda_margin": 0.5576,
        "net_income": 29320, "eps": 6.01,
        "comment": "Source: Broadcom Q2 FY2026 earnings release, accessed 2026-06-25. https://investors.broadcom.com",
    },
    {
        "name": "Qualcomm", "ticker": "QCOM",
        "revenue": 44487, "rev_growth": 0.052,
        "gross_margin": 0.5480, "ebitda": 13000, "ebitda_margin": 0.2922,
        "net_income": 9920, "eps": 9.15,
        "comment": "Source: Qualcomm Q2 FY2026 10-Q filing, SEC EDGAR, accessed 2026-06-25. https://investor.qualcomm.com",
    },
    {
        "name": "Intel", "ticker": "INTC",
        "revenue": 53763, "rev_growth": 0.014,
        "gross_margin": 0.3720, "ebitda": 14170, "ebitda_margin": 0.2635,
        "net_income": -3170, "eps": -0.67,
        "comment": "Source: Intel Q1 FY2026 earnings release, accessed 2026-06-25. https://www.intc.com. Negative net income reflects foundry investments and restructuring.",
    },
    {
        "name": "Micron", "ticker": "MU",
        "revenue": 90270, "rev_growth": 0.770,
        "gross_margin": 0.7257, "ebitda": 68220, "ebitda_margin": 0.7557,
        "net_income": 50470, "eps": 44.31,
        "comment": "Source: Micron Q2 FY2026 earnings release, accessed 2026-06-25. Record revenue driven by AI/HBM memory demand surge. https://investors.micron.com",
    },
]

for i, co in enumerate(companies):
    r = 7 + i
    ws.cell(row=r, column=1, value=co["name"]).font = font_company
    ws.cell(row=r, column=1).alignment = center_align

    # Revenue (B)
    b = ws.cell(row=r, column=2, value=co["revenue"])
    b.font = font_input
    b.number_format = '#,##0'
    b.alignment = center_align
    b.comment = Comment(co["comment"], "Comps Analyst", width=350, height=120)

    # Revenue Growth (C)
    c = ws.cell(row=r, column=3, value=co["rev_growth"])
    c.font = font_input
    c.number_format = '0.0%'
    c.alignment = center_align

    # Gross Profit (D) = Revenue * Gross Margin
    d = ws.cell(row=r, column=4)
    d.value = f"=B{r}*E{r}"
    d.font = font_data
    d.number_format = '#,##0'
    d.alignment = center_align

    # Gross Margin (E)
    e = ws.cell(row=r, column=5, value=co["gross_margin"])
    e.font = font_input
    e.number_format = '0.0%'
    e.alignment = center_align

    # EBITDA (F)
    f_cell = ws.cell(row=r, column=6, value=co["ebitda"])
    f_cell.font = font_input
    f_cell.number_format = '#,##0'
    f_cell.alignment = center_align

    # EBITDA Margin (G) = EBITDA / Revenue
    g = ws.cell(row=r, column=7)
    g.value = f"=F{r}/B{r}"
    g.font = font_data
    g.number_format = '0.0%'
    g.alignment = center_align

    # Net Income (H)
    h = ws.cell(row=r, column=8, value=co["net_income"])
    h.font = font_input
    h.number_format = '#,##0'
    h.alignment = center_align

    # EPS (I)
    i_cell = ws.cell(row=r, column=9, value=co["eps"])
    i_cell.font = font_input
    i_cell.number_format = '#,##0.00'
    i_cell.alignment = center_align

    for col in range(1, 10):
        ws.cell(row=r, column=col).fill = white_fill

for r in range(7, 13):
    ws.row_dimensions[r].height = 22

# ============================================================
# STATISTICS ROWS (Rows 14-18)
# ============================================================
stats_labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
stat_funcs = [
    lambda col: f"=MAX({col}7:{col}12)",
    lambda col: f"=QUARTILE({col}7:{col}12,3)",
    lambda col: f"=MEDIAN({col}7:{col}12)",
    lambda col: f"=QUARTILE({col}7:{col}12,1)",
    lambda col: f"=MIN({col}7:{col}12)",
]

for si, (label, func) in enumerate(zip(stats_labels, stat_funcs)):
    r = 14 + si
    ws.cell(row=r, column=1, value=label).font = font_stat_label
    ws.cell(row=r, column=1).fill = light_grey_fill
    ws.cell(row=r, column=1).alignment = left_align

    for col_idx in range(2, 10):
        letter = get_column_letter(col_idx)
        cell = ws.cell(row=r, column=col_idx)
        cell.value = func(letter)
        cell.font = font_data
        cell.fill = light_grey_fill
        cell.alignment = center_align
        if col_idx in [3, 5, 7]:
            cell.number_format = '0.0%'
        elif col_idx == 9:
            cell.number_format = '#,##0.00'
        elif col_idx in [2, 4, 6, 8]:
            cell.number_format = '#,##0'

print("Operating section complete.")

# ============================================================
# ROW 20: SECTION HEADER — VALUATION MULTIPLES
# ============================================================
row = 20
ws.merge_cells(f"A{row}:I{row}")
ws[f"A{row}"] = "VALUATION MULTIPLES & INVESTMENT METRICS"
ws[f"A{row}"].font = font_section
ws[f"A{row}"].fill = dark_blue_fill
ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 10):
    ws.cell(row=row, column=col).fill = dark_blue_fill

# ============================================================
# ROW 21: COLUMN HEADERS
# ============================================================
row = 21
v_headers = [
    "Company", "Market Cap", "Enterprise\nValue", "EV / Revenue\n(LTM)",
    "EV / EBITDA\n(LTM)", "P / E Ratio", "Beta\n(5Y Monthly)", "Dividend\nYield", "FCF Yield"
]
for col_idx, header in enumerate(v_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = font_col_header
    cell.fill = light_blue_fill
    cell.alignment = center_align
ws.row_dimensions[row].height = 35

# ============================================================
# VALUATION DATA (Rows 22-27)
# ============================================================
valuation_data = [
    {
        "name": "AMD",
        "mkt_cap": 847490, "ev": 839010,
        "beta": 2.49, "div_yield": None,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25.",
    },
    {
        "name": "NVIDIA",
        "mkt_cap": 4820000, "ev": 4780000,
        "beta": 2.20, "div_yield": 0.0050,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25. https://finance.yahoo.com/quote/NVDA",
    },
    {
        "name": "Broadcom",
        "mkt_cap": 1820000, "ev": 1860000,
        "beta": 1.43, "div_yield": 0.0068,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25. https://finance.yahoo.com/quote/AVGO",
    },
    {
        "name": "Qualcomm",
        "mkt_cap": 208070, "ev": 212440,
        "beta": 1.60, "div_yield": 0.0186,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25. https://finance.yahoo.com/quote/QCOM",
    },
    {
        "name": "Intel",
        "mkt_cap": 661670, "ev": 673600,
        "beta": 2.23, "div_yield": None,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25. Dividend suspended.",
    },
    {
        "name": "Micron",
        "mkt_cap": 1180000, "ev": 1160000,
        "beta": 2.17, "div_yield": 0.0006,
        "comment": "Source: Yahoo Finance / Stock Analysis, market data as of 2026-06-25. https://finance.yahoo.com/quote/MU",
    },
]

for i, co in enumerate(valuation_data):
    r = 22 + i
    op_row = 7 + i

    ws.cell(row=r, column=1, value=co["name"]).font = font_company
    ws.cell(row=r, column=1).alignment = center_align

    # Market Cap (B)
    b = ws.cell(row=r, column=2, value=co["mkt_cap"])
    b.font = font_input
    b.number_format = '#,##0'
    b.alignment = center_align
    b.comment = Comment(co["comment"], "Comps Analyst", width=350, height=80)

    # Enterprise Value (C)
    c = ws.cell(row=r, column=3, value=co["ev"])
    c.font = font_input
    c.number_format = '#,##0'
    c.alignment = center_align

    # EV/Revenue (D) = references Operating section B column
    d = ws.cell(row=r, column=4)
    d.value = f"=C{r}/B{op_row}"
    d.font = font_data
    d.number_format = '0.0"x"'
    d.alignment = center_align

    # EV/EBITDA (E) = references Operating section F column
    e = ws.cell(row=r, column=5)
    e.value = f"=C{r}/F{op_row}"
    e.font = font_data
    e.number_format = '0.0"x"'
    e.alignment = center_align

    # P/E (F)
    f_cell = ws.cell(row=r, column=6)
    f_cell.value = f'=IF(H{op_row}>0,B{r}/H{op_row},"N/M")'
    f_cell.font = font_data
    f_cell.number_format = '0.0"x"'
    f_cell.alignment = center_align

    # Beta (G)
    g = ws.cell(row=r, column=7, value=co["beta"])
    g.font = font_input
    g.number_format = '0.00'
    g.alignment = center_align

    # Dividend Yield (H)
    h = ws.cell(row=r, column=8)
    if co["div_yield"] is not None:
        h.value = co["div_yield"]
        h.font = font_input
        h.number_format = '0.00%'
    else:
        h.value = "N/A"
        h.font = font_data
    h.alignment = center_align

    # FCF Yield (I)
    i_cell = ws.cell(row=r, column=9)
    i_cell.value = "N/A"
    i_cell.font = font_data
    i_cell.alignment = center_align

    for col in range(1, 10):
        ws.cell(row=r, column=col).fill = white_fill

# ============================================================
# VALUATION STATISTICS (Rows 29-33)
# ============================================================
for si, (label, func) in enumerate(zip(stats_labels, stat_funcs)):
    r = 29 + si
    ws.cell(row=r, column=1, value=label).font = font_stat_label
    ws.cell(row=r, column=1).fill = light_grey_fill
    ws.cell(row=r, column=1).alignment = left_align

    for col_idx in range(2, 10):
        letter = get_column_letter(col_idx)
        cell = ws.cell(row=r, column=col_idx)
        cell.value = func(letter)
        cell.font = font_data
        cell.fill = light_grey_fill
        cell.alignment = center_align
        if col_idx == 7:
            cell.number_format = '0.00'
        elif col_idx == 8:
            cell.number_format = '0.00%'
        elif col_idx in [4, 5, 6]:
            cell.number_format = '0.0"x"'
        elif col_idx in [2, 3]:
            cell.number_format = '#,##0'

# ============================================================
# NOTES & METHODOLOGY SECTION
# ============================================================
notes_start = 35
ws.merge_cells(f"A{notes_start}:I{notes_start}")
ws[f"A{notes_start}"] = "NOTES & METHODOLOGY"
ws[f"A{notes_start}"].font = font_section
ws[f"A{notes_start}"].fill = dark_blue_fill
ws[f"A{notes_start}"].alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 10):
    ws.cell(row=notes_start, column=col).fill = dark_blue_fill

notes = [
    ("Data Sources:", "All financial data sourced from company SEC filings (10-K, 10-Q), earnings releases, and market data providers (Yahoo Finance, Stock Analysis). Data accessed June 25, 2026."),
    ("Time Period:", "LTM (Last Twelve Months) figures as of the most recent quarterly filing for each company. Market data as of market close June 24-25, 2026."),
    ("EBITDA Definition:", "EBITDA = Operating Income + Depreciation & Amortization. Where not directly disclosed, estimated from operating income and D&A from cash flow statements."),
    ("Enterprise Value:", "EV = Market Capitalization + Total Debt - Cash & Cash Equivalents. Reflects net debt position as of the most recent balance sheet date."),
    ("P/E Ratio:", "P/E = Market Cap / Net Income (LTM). Displayed as N/M (not meaningful) for companies with negative net income (e.g., Intel)."),
    ("Revenue Growth:", "Year-over-year percentage change comparing LTM revenue to the prior-year comparable period."),
    ("Key Observations:", ""),
    ("  * NVIDIA:", "Leads in scale ($253B TTM revenue), profitability (74% gross margin, 65% EBITDA margin), and growth (85% YoY). Dominates AI GPU market."),
    ("  * Micron:", "Highest growth rate (~77% YoY) driven by AI/HBM memory cycle upswing. Memory is historically cyclical."),
    ("  * Broadcom:", "Strong software-like margins (76% gross) post-VMware integration. Custom AI chip revenue growing rapidly."),
    ("  * AMD:", "High P/E (173x) and EV/EBITDA (113x) reflect market pricing in significant future AI data center growth. Verify EBITDA figures."),
    ("  * Intel:", "Negative net income makes P/E not meaningful. EV/EBITDA of 48x may overstate value given margin compression (37% gross margin)."),
    ("  * Qualcomm:", "Highest dividend yield (1.86%) but slowest growth (+5% YoY). Mobile exposure creates cyclicality risk."),
    ("Red Flags:", ""),
    ("  * AMD:", "EV/EBITDA of 113x is extremely elevated. Verify with latest 10-Q; may reflect one-time items or accounting treatment differences in EBITDA."),
    ("  * Intel:", "EV/EBITDA of 48x appears stretched for a company with negative earnings and compressed margins. Foundry ramp uncertainty persists."),
    ("  * Micron:", "77% revenue growth and 76% gross margin reflect cyclical peak. Memory is historically volatile; normalize before valuation."),
    ("Cross-Check:", "All LTM figures should be independently verified against latest 10-Q/10-K filings before making investment decisions."),
    ("Disclaimer:", "This analysis is for informational purposes only and does not constitute investment advice."),
]

for i, (label, text) in enumerate(notes):
    r = notes_start + 1 + i
    if label:
        ws.cell(row=r, column=1, value=label).font = Font(name=font_tnr, size=11, bold=True)
        ws.cell(row=r, column=1).fill = light_blue_fill
        for col in range(1, 10):
            ws.cell(row=r, column=col).fill = light_blue_fill
    if text:
        ws.merge_cells(f"A{r}:I{r}")
        ws.cell(row=r, column=1, value=text).font = Font(name=font_tnr, size=10)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 20

# ============================================================
# SAVE
# ============================================================
output_dir = project_path(config_value("files", "output_dir", "data/output"))
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "Semiconductor_Comps_Analysis.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print("Done!")
