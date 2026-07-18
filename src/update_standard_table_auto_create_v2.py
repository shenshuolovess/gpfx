# -*- coding: utf-8 -*-
"""
将每次分析生成的“分类总表”写入“标准表”：
1）标准表不存在：第一次运行自动新建；
2）标准表已存在：在原标准表基础上更新；
3）新增股票只按“代码”判断；代码不存在于标准表才新增行；
4）每次只新增/更新一个日期列；
5）标准表已有股票但本次分析文件不存在：本次日期列置空；
6）分析文件有新股票但标准表没有：自动新增行；
7）更新后按每行日期列非空数量从多到少排序；
8）非空数量相同时，依次按 赶顶数量、上升数量、震荡上行数量 从多到少排序；
9）对所有日期列都不为空的股票，把该行日期分类单元格涂红；
10）输出为 xlsx 时，会设置列宽、冻结首行、自动筛选；CSV 不支持保存列宽和颜色；
11）默认覆盖前生成 .bak_时间戳 备份。

推荐运行方式：
python src/update_standard_table_auto_create_v2.py

不备份运行：
python src/update_standard_table_auto_create_v2.py --no-backup
"""

import argparse
import os
import re
import shutil
from datetime import datetime
from typing import Optional

import pandas as pd

from pipeline_config import config_value, project_path, resolve_input
from stock_utils import (
    normalize_code_series,
    read_table,
    write_csv,
)

BASE_COLS = ["代码", "名称", "市场"]
REQUIRED_ANALYSIS_COLS = ["代码", "名称", "市场", "截止交易日", "分类"]

# 列宽设置
DEFAULT_DATE_COL_WIDTH = 18
DEFAULT_CODE_COL_WIDTH = 14
DEFAULT_NAME_COL_WIDTH = 18
DEFAULT_MARKET_COL_WIDTH = 10

# 排序优先级：非空数量相同后，按下面三个分类数量依次比较
PRIORITY_CLASSES = ["赶顶", "上升", "震荡上行"]

# 对“每个日期列都有值”的股票，涂红日期分类单元格
FULL_ROW_RED_FILL = "FFC7CE"   # 浅红底
FULL_ROW_RED_FONT = "9C0006"   # 深红字


def get_file_ext(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def format_date_col(date_value: str) -> str:
    """统一日期列格式为 2026/5/15。"""
    if pd.isna(date_value):
        raise ValueError("截止交易日为空，无法生成日期列")

    text = str(date_value).strip()

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            dt = datetime.strptime(text, fmt)
            return f"{dt.year}/{dt.month}/{dt.day}"
        except ValueError:
            pass

    m = re.search(r"(20\d{6})", text)
    if m:
        dt = datetime.strptime(m.group(1), "%Y%m%d")
        return f"{dt.year}/{dt.month}/{dt.day}"

    raise ValueError(f"无法识别日期格式：{date_value}")


def infer_date_col(analysis_df: pd.DataFrame, analysis_file: str) -> str:
    """优先从分析文件的 截止交易日 列取日期；取不到再从文件名取。"""
    if "截止交易日" in analysis_df.columns:
        dates = analysis_df["截止交易日"].dropna().astype(str).str.strip()
        dates = dates[dates != ""]
        if not dates.empty:
            return format_date_col(dates.mode().iloc[0])

    m = re.search(r"(20\d{6})", os.path.basename(analysis_file))
    if m:
        return format_date_col(m.group(1))

    raise ValueError("无法从 截止交易日 列或文件名推断本次日期")


def validate_columns(df: pd.DataFrame, required_cols: list[str], file_name: str) -> None:
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"{file_name} 缺少必要字段：{missing}，当前字段：{list(df.columns)}")


def build_analysis_table(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """从分析文件中提取标准表需要的字段，重复代码保留最后一条。"""
    validate_columns(analysis_df, REQUIRED_ANALYSIS_COLS, "分析文件")

    tmp = analysis_df[BASE_COLS + ["分类"]].copy()
    tmp["代码"] = normalize_code_series(tmp["代码"], style="suffix")
    tmp["名称"] = tmp["名称"].fillna("").astype(str).str.strip()
    tmp["市场"] = tmp["市场"].fillna("").astype(str).str.strip()
    tmp["分类"] = tmp["分类"].fillna("").astype(str).str.strip()

    tmp = tmp[tmp["代码"] != ""].copy()

    duplicated_rows = int(tmp.duplicated(subset=["代码"], keep=False).sum())
    if duplicated_rows > 0:
        print(f"[WARN] 分析文件存在重复代码记录 {duplicated_rows} 行，将按代码保留最后一条")
        tmp = tmp.drop_duplicates(subset=["代码"], keep="last")

    return tmp.reset_index(drop=True)


def read_or_create_standard_table(standard_file: str) -> pd.DataFrame:
    """标准表存在则读取；不存在则创建空标准表。"""
    if os.path.exists(standard_file):
        standard_df = read_table(standard_file, dtype=str)
        validate_columns(standard_df, BASE_COLS, "标准表")
        print(f"[INFO] 已读取标准表：{standard_file}，原始行数：{len(standard_df)}")
        return standard_df

    print(f"[INFO] 标准表不存在，第一次运行将自动新建：{standard_file}")
    return pd.DataFrame(columns=BASE_COLS)


def get_date_cols(df: pd.DataFrame) -> list[str]:
    """标准表中除 代码/名称/市场 外，其他列全部视为日期结果列。"""
    return [c for c in df.columns if c not in BASE_COLS]


def sort_by_non_empty_and_priority(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    排序规则：
    1）每行日期列非空数量多的在前；
    2）非空数量相同，赶顶数量多的在前；
    3）再相同，上升数量多的在前；
    4）再相同，震荡上行数量多的在前；
    5）最后保持原相对顺序。
    """
    result = df.copy()
    date_cols = get_date_cols(result)

    if not date_cols:
        return result, {
            "date_col_count": 0,
            "full_non_empty_rows": 0,
            "max_non_empty_count": 0,
        }

    date_part = result[date_cols].fillna("").astype(str).apply(lambda col: col.str.strip())

    result["__原顺序__"] = range(len(result))
    result["__非空数量__"] = date_part.ne("").sum(axis=1)

    for cls in PRIORITY_CLASSES:
        result[f"__{cls}数量__"] = date_part.eq(cls).sum(axis=1)

    sort_cols = ["__非空数量__"] + [f"__{cls}数量__" for cls in PRIORITY_CLASSES] + ["__原顺序__"]
    ascending = [False] * (1 + len(PRIORITY_CLASSES)) + [True]

    result = result.sort_values(sort_cols, ascending=ascending, kind="mergesort").reset_index(drop=True)

    helper_cols = ["__原顺序__", "__非空数量__"] + [f"__{cls}数量__" for cls in PRIORITY_CLASSES]
    stats = {
        "date_col_count": len(date_cols),
        "full_non_empty_rows": int((date_part.ne("").sum(axis=1) == len(date_cols)).sum()),
        "max_non_empty_count": int(date_part.ne("").sum(axis=1).max()) if len(result) else 0,
    }

    result = result.drop(columns=helper_cols)
    return result, stats


def write_table_with_format(df: pd.DataFrame, output_file: str, current_date_col: str) -> None:
    """写出CSV或XLSX；XLSX支持设置列宽和涂色，CSV不支持。"""
    ext = get_file_ext(output_file)

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("输出xlsx需要安装openpyxl：pip install openpyxl") from e

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="标准表")

        wb = load_workbook(output_file)
        ws = wb["标准表"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        full_row_fill = PatternFill("solid", fgColor=FULL_ROW_RED_FILL)
        full_row_font = Font(color=FULL_ROW_RED_FONT)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        width_map = {
            "代码": DEFAULT_CODE_COL_WIDTH,
            "名称": DEFAULT_NAME_COL_WIDTH,
            "市场": DEFAULT_MARKET_COL_WIDTH,
        }

        date_col_indexes = []
        for idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(idx)
            if col_name in BASE_COLS:
                ws.column_dimensions[letter].width = width_map.get(col_name, 12)
            else:
                # 日期列统一稍宽一些。
                ws.column_dimensions[letter].width = DEFAULT_DATE_COL_WIDTH
                date_col_indexes.append(idx)

        # 对齐所有数据单元格。
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = center

        # 对“所有日期列都不为空”的股票，把该行所有日期分类单元格涂红。
        # 注意：只涂日期结果列，不涂 代码/名称/市场。
        if date_col_indexes:
            for row_idx in range(2, ws.max_row + 1):
                date_cells = [ws.cell(row=row_idx, column=col_idx) for col_idx in date_col_indexes]
                all_date_not_empty = all(str(cell.value or "").strip() != "" for cell in date_cells)
                if all_date_not_empty:
                    for cell in date_cells:
                        cell.fill = full_row_fill
                        cell.font = full_row_font

        wb.save(output_file)
        return

    if ext == ".csv" or ext == "":
        write_csv(df, output_file)
        print("[WARN] 当前输出为CSV，CSV文件本身无法保存列宽和颜色；如需日期列变宽、单元格涂红，请使用 XLSX 标准表")
        return

    raise ValueError(f"不支持的输出格式：{output_file}，请使用 .csv 或 .xlsx")


def update_standard_table(
    analysis_file: str,
    standard_file: str,
    output_file: Optional[str] = None,
    backup: bool = True,
) -> str:
    if output_file is None:
        output_file = standard_file

    analysis_df = read_table(analysis_file, dtype=str)
    date_col = infer_date_col(analysis_df, analysis_file)
    analysis_small = build_analysis_table(analysis_df)

    standard_df = read_or_create_standard_table(standard_file)
    result_df = standard_df.copy()

    # 补齐基础列，避免第一次空表或历史标准表缺列导致异常。
    for col in BASE_COLS:
        if col not in result_df.columns:
            result_df[col] = ""

    result_df["代码"] = normalize_code_series(result_df["代码"], style="suffix")
    result_df["名称"] = result_df["名称"].fillna("").astype(str).str.strip()
    result_df["市场"] = result_df["市场"].fillna("").astype(str).str.strip()

    # 标准表内部若有空代码或重复代码，严格按代码清理，保留最后一条。
    before_dedup = len(result_df)
    result_df = result_df[result_df["代码"] != ""].copy()
    result_df = result_df.drop_duplicates(subset=["代码"], keep="last").reset_index(drop=True)
    if len(result_df) != before_dedup:
        print(f"[WARN] 标准表存在空代码或重复代码，已按代码清理：{before_dedup} -> {len(result_df)}")

    # 显式说明：后续所有新增判断只看代码，不看名称/市场。
    existing_codes = set(result_df["代码"].astype(str).str.strip())
    existing_codes.discard("")

    class_map = dict(zip(analysis_small["代码"], analysis_small["分类"]))
    name_map = dict(zip(analysis_small["代码"], analysis_small["名称"]))
    market_map = dict(zip(analysis_small["代码"], analysis_small["市场"]))

    # 旧股票：写入本次日期列；本次分析没有的置空。
    result_df[date_col] = result_df["代码"].map(class_map).fillna("")

    # 已有股票只在名称/市场为空时补齐，不因为名称或市场变化新增行。
    for col, mp in (("名称", name_map), ("市场", market_map)):
        empty_mask = result_df[col].astype(str).str.strip().eq("")
        result_df.loc[empty_mask, col] = result_df.loc[empty_mask, "代码"].map(mp).fillna("")

    # 新股票：只按代码判断。分析文件代码不在标准表中，才追加。
    new_rows = analysis_small[~analysis_small["代码"].isin(existing_codes)].copy()
    new_count = len(new_rows)

    if new_count > 0:
        append_df = pd.DataFrame("", index=range(new_count), columns=result_df.columns)
        append_df["代码"] = new_rows["代码"].values
        append_df["名称"] = new_rows["名称"].values
        append_df["市场"] = new_rows["市场"].values
        append_df[date_col] = new_rows["分类"].values
        result_df = pd.concat([result_df, append_df], ignore_index=True)

    # 保持原标准表列顺序，只把本次日期列放在最后；第一次新建则为 代码/名称/市场/日期。
    cols_without_date = [c for c in result_df.columns if c != date_col]
    ordered_cols = [c for c in cols_without_date if c in result_df.columns] + [date_col]
    result_df = result_df[ordered_cols].fillna("")

    # 更新完成后排序：按非空数量、赶顶数量、上升数量、震荡上行数量排序。
    result_df, sort_stats = sort_by_non_empty_and_priority(result_df)

    # 只有覆盖已存在文件时才备份；第一次创建没有原文件，不备份。
    if backup and os.path.exists(output_file):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"{output_file}.bak_{ts}"
        shutil.copy2(output_file, backup_file)
        print(f"[INFO] 已备份原标准表：{backup_file}")

    write_table_with_format(result_df, output_file, date_col)

    matched_count = int(result_df[date_col].astype(str).str.strip().ne("").sum())
    empty_count = int(len(result_df) - matched_count)

    print("[INFO] 标准表更新完成")
    print(f"[INFO] 分析文件：{analysis_file}")
    print(f"[INFO] 标准表：{standard_file}")
    print(f"[INFO] 输出文件：{output_file}")
    print(f"[INFO] 本次日期列：{date_col}")
    print("[INFO] 新增股票判断口径：代码")
    print(f"[INFO] 分析文件有效股票数：{len(analysis_small)}")
    print(f"[INFO] 本次新增股票行数：{new_count}")
    print(f"[INFO] 本次日期列有值：{matched_count}")
    print(f"[INFO] 本次日期列为空：{empty_count}")
    print(f"[INFO] 参与排序的日期列数量：{sort_stats['date_col_count']}")
    print(f"[INFO] 单行最大非空日期数量：{sort_stats['max_non_empty_count']}")
    print(f"[INFO] 所有日期列都不为空的股票数：{sort_stats['full_non_empty_rows']}")
    print(f"[INFO] 输出标准表总行数：{len(result_df)}")
    print(f"[INFO] 输出列顺序：{list(result_df.columns)}")

    return output_file


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="把分类总表结果写入标准表；标准表不存在则第一次自动新建；新增股票按代码判断；按非空数量和分类优先级排序；xlsx可涂红"
    )
    parser.add_argument("--analysis", help="分类总表 CSV/XLSX；默认自动选择最新 CSV")
    parser.add_argument(
        "--standard",
        default=config_value("files", "standard_table", "data/input/标准表.xlsx"),
        help="标准表 CSV/XLSX；默认读取统一配置",
    )
    parser.add_argument("--output", default=None, help="输出文件；不填则覆盖/创建standard")
    parser.add_argument("--no-backup", action="store_true", help="覆盖输出时不备份原标准表")
    args = parser.parse_args(argv)

    analysis_file = resolve_input(args.analysis, pattern_key="classification_pattern")
    standard_file = project_path(args.standard)
    output_file = project_path(args.output) if args.output else None

    update_standard_table(
        analysis_file=str(analysis_file),
        standard_file=str(standard_file),
        output_file=str(output_file) if output_file else None,
        backup=not args.no_backup,
    )


if __name__ == "__main__":
    main()
